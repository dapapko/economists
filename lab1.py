import itertools
import statistics
import openpyxl
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np


# Wrapper which allows us to parse data and make queries to the parsed information
class ExcelDB:
    def __init__(self, db=None, file=None):
        if db is None:
            db = []
        self.db = db
        if file is not None:
            self.from_file(file)

# Excel parsing
    def from_file(self, path):
        workbook = openpyxl.load_workbook(path)
        ws = workbook.active
        for row in ws.iter_rows(min_row=5, max_row=101, min_col=2, max_col=19):
            for cell in row:
                if cell.value is not None:
                    entry = {"year": int(ws[f'{chr(cell.column + 64)}4'].value),
                             'region': ws[f'A{cell.row}'].value,
                             'salary': int(cell.value)}
                    self.db.append(entry)

# Here is just a set of fiters used for queries

    def where(self, field, value, negation=False):
        filter = [x for x in self.db if x[field] == value] if not negation else [x for x in self.db if
                                                                                 not x[field] == value]
        return ExcelDB(filter)

    def where_in(self, field, value, negation=False):
        filter = [
            x for x in self.db if x[field] in value
        ] if not negation else [
            x for x in self.db if not x[field] not in value]
        return ExcelDB(filter)

    def where_between(self, field, inf, sup, negation=False):
        filter = [x for x in self.db if inf <= x[field] <= sup] if not negation else [x for x in self.db if
                                                                                      not inf <= x[field] <= sup]
        return ExcelDB(filter)

    def where_less_or_equal(self, field, value, negation=False):
        filter = [x for x in self.db if x[field] <= value] if not negation else [x for x in self.db if
                                                                                 not x[field] <= value]
        return ExcelDB(filter)

    def where_more_or_equal(self, field, value, negation=False):
        filter = [x for x in self.db if value <= x[field]] if not negation else [x for x in self.db if
                                                                                 not value <= x[field]]
        return ExcelDB(filter)

    def pluck(self, field: str) -> tuple:
        return tuple([x[field] for x in self.db])


# Disjunction (logical OR) for queries (union of multiple queries)
    @staticmethod
    def disjunction(*args):
        db = []
        for a in args:
            db.extend(a.db)
        return ExcelDB(db)

# Calculate average increase of some numeric parameter by finding deltas

    def avg_increase(self, field: str):
        if len(self.db) == 0:
            raise ValueError
        elif len(self.db) == 1:
            return self.db[0][field]
        slowpoke, leader = itertools.tee(self.db)
        leader.__next__()
        increases = []
        for prev, next in zip(slowpoke, leader):
            increases.append(next[field] - prev[field])
        return statistics.mean(increases)

# Some useful statistical methods

    def max(self, field: str):
        return max([x[field] for x in self.db])

    def min(self, field: str):
        return min([x[field] for x in self.db])

    def mean(self, field: str):
        return statistics.mean([x[field] for x in self.db])

    def median(self, field: str) -> float:
        return statistics.median([x[field] for x in self.db])

    def mode(self, field: str) -> float:
        return statistics.mode([x[field] for x in self.db])


xlsx_file = Path('C:\stat', 'statistika.xlsx')
# Init database from file
db = ExcelDB(file=xlsx_file)
# Making queries
nn = db.where('region', 'Нижегородская область').where_between('year', 2011, 2015).pluck('salary')
vl = db.where('region', 'Владимирская область').where_between('year', 2011, 2015).pluck('salary')

# Draw a line and a bar chart on the same page
fig, axs = plt.subplots(2,1)
ind = np.arange(5)
width = 0.35
b1 = axs[0].bar(ind, nn, width, label='Нижний Новгород')
b2 = axs[0].bar(ind + width, vl, width, label='Владимир')
years = ['2011', '2012', '2013', '2014', '2015']
axs[0].set_title('Зарплата во Владимире и НН  с 2011 по 2015 год')
axs[0].set_xticks([r + width for r in range(len(vl))])
axs[0].set_xticklabels(years)
axs[0].set_ylabel('Заработная плата (в рублях)', fontweight='bold')
axs[0].legend((b1, b2), ('Нижний Новгород', 'Владимир'))
axs[1].set_title('Сравнение ЗП во Владимирской и Московской области (2011-2015)', fontsize=11)
msk = db.where_between('year', 2011, 2015).where('region', 'Московская область').pluck('salary')
axs[1].plot(years, msk, 'ro')
axs[1].plot(years, vl, 'bo')
plt.show()

# Draw a pie chart on the separate page
years = ['2011', '2012', '2013', '2014', '2015']
plt.figure(figsize=plt.figaspect(1))
plt.title('Зарплата во Владимире с 2011 по 2015 год')
plt.pie(vl, labels=years, autopct=lambda p: f'{(sum(vl) / 100) * p:.0f} ({p:.2f}%)')
plt.show()


